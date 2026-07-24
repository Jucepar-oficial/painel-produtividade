#!/usr/bin/env python3
"""Gera arquivos JSON agregados a partir da planilha de produtividade da Jucepar."""

import argparse
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


def texto(valor):
    return str(valor or "").strip()


def carregar_entidades(wb):
    if "Tabelas" not in wb.sheetnames:
        raise ValueError("A aba 'Tabelas' não foi encontrada.")
    entidades = {}
    for linha in wb["Tabelas"].iter_rows(min_row=2, values_only=True):
        nome = texto(linha[0]).upper()
        if nome:
            entidades[nome] = {
                "entidade": texto(linha[1]),
                "sigla": texto(linha[2]) or texto(linha[1]),
            }
    if not entidades:
        raise ValueError("Nenhum vogal foi encontrado na aba 'Tabelas'.")
    return entidades


def gerar(arquivo, competencia, saida):
    partes = competencia.split("-")
    try:
        ano = int(partes[0])
        mes = int(partes[1]) if len(partes) == 2 else None
    except Exception as exc:
        raise ValueError("Use AAAA para um ano ou AAAA-MM para um mês.") from exc
    if len(partes) not in (1, 2) or (mes is not None and not 1 <= mes <= 12):
        raise ValueError("Use AAAA para um ano ou AAAA-MM para um mês.")

    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True, keep_links=False)
    if "DADOS" not in wb.sheetnames:
        raise ValueError("A aba 'DADOS' não foi encontrada.")
    entidades = carregar_entidades(wb)
    grupos = defaultdict(int)
    linhas_lidas = 0
    meses_encontrados = set()

    for linha in wb["DADOS"].iter_rows(min_row=2, values_only=True):
        nome, tipo, data, _protocolo, decisao = linha[:5]
        if not data:
            continue
        if not hasattr(data, "year"):
            raise ValueError(f"Data inválida encontrada: {data!r}")
        meses_encontrados.add(f"{data.year:04d}-{data.month:02d}")
        if data.year != ano or (mes is not None and data.month != mes):
            continue
        nome = texto(nome).upper()
        if nome not in entidades:
            raise ValueError(f"Vogal sem cadastro na aba 'Tabelas': {nome}")
        periodo = f"{data.year:04d}-{data.month:02d}"
        grupos[(nome, periodo, texto(tipo), texto(decisao))] += 1
        linhas_lidas += 1

    if not linhas_lidas:
        encontrados = ", ".join(sorted(meses_encontrados)) or "nenhum"
        raise ValueError(f"Não há registros para {competencia}. Competências encontradas: {encontrados}.")

    registros = []
    for (nome, periodo, tipo, decisao), quantidade in sorted(grupos.items()):
        cadastro = entidades[nome]
        registros.append({
            "vogal": nome,
            "periodo": periodo,
            "tipoAnalise": tipo,
            "decisao": decisao,
            "entidade": cadastro["entidade"],
            "siglaEntidade": cadastro["sigla"],
            "quantidade": quantidade,
        })

    payload = {
        "competencia": competencia,
        "geradoEm": datetime.now().astimezone().isoformat(timespec="seconds"),
        "totalRegistros": linhas_lidas,
        "totalCombinacoes": len(registros),
        "registros": registros,
    }
    saida.parent.mkdir(parents=True, exist_ok=True)
    saida.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    return payload


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("planilha", type=Path)
    parser.add_argument("competencia", help="AAAA ou AAAA-MM")
    parser.add_argument("saida", type=Path)
    args = parser.parse_args()
    payload = gerar(args.planilha, args.competencia, args.saida)
    print(f"{payload['competencia']}: {payload['totalRegistros']} registros em {payload['totalCombinacoes']} combinações")


if __name__ == "__main__":
    main()
